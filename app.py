import os
import re
import tempfile
from flask import Flask, render_template, request, send_from_directory, redirect, url_for
import pdfplumber
from pdf2image import convert_from_path
import pytesseract
from openpyxl import load_workbook

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

MAPPING = {
    '姓名': 'B2',
    '身份证号': 'C2',
    '单位名称': 'D2'
}

def extract_text_from_pdf(pdf_path):
    text = ''
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ''
            text += page_text
    if text.strip():
        return text
    images = convert_from_path(pdf_path)
    for img in images:
        text += pytesseract.image_to_string(img, lang='chi_sim')
    return text

def parse_fields(text):
    fields = {}
    for key in MAPPING:
        match = re.search(rf'{key}[：:]*\s*(\S+)', text)
        if match:
            fields[key] = match.group(1)
    return fields

def fill_excel(template_path, fields):
    wb = load_workbook(template_path)
    ws = wb.active
    for key, cell in MAPPING.items():
        if key in fields:
            ws[cell] = fields[key]
    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(tmp_fd)
    wb.save(tmp_path)
    return tmp_path

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        pdf_files = request.files.getlist('pdfs')
        template = request.files['template']
        template_path = os.path.join(app.config['UPLOAD_FOLDER'], template.filename)
        template.save(template_path)

        combined_text = ''
        for f in pdf_files:
            pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], f.filename)
            f.save(pdf_path)
            combined_text += extract_text_from_pdf(pdf_path)

        fields = parse_fields(combined_text)
        output_path = fill_excel(template_path, fields)
        filename = os.path.basename(output_path)
        return redirect(url_for('download_file', filename=filename))
    return render_template('index.html')

@app.route('/downloads/<filename>')
def download_file(filename):
    return send_from_directory('.', filename, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
